import datetime
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse_lazy, reverse
from django.contrib import messages
from django.views.generic import TemplateView, ListView, CreateView, UpdateView, DeleteView, View
from django.contrib.auth.views import LoginView, LogoutView, PasswordChangeView
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.contrib.messages.views import SuccessMessageMixin
from django.db.models import Sum, Count, Avg, Q
from django.http import HttpResponse, Http404

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

from django.contrib.auth.models import User
from .models import Course, Teacher, Subject, Student, Attendance, Result, Fee, Notice, Timetable
from .forms import (
    TeacherUserForm, TeacherProfileForm, StudentForm, CourseForm, 
    SubjectForm, AttendanceForm, ResultForm, FeeForm, NoticeForm, TimetableForm,
    CampusPasswordChangeForm
)

from django.contrib.auth.decorators import user_passes_test

def admin_required(view_func):
    actual_decorator = user_passes_test(
        lambda u: u.is_active and u.is_superuser,
        login_url='login'
    )
    return actual_decorator(view_func)

# ==========================================
# ACCESS CONTROL MIXINS
# ==========================================

class AdminRequiredMixin(LoginRequiredMixin, UserPassesTestMixin):
    def test_func(self):
        return self.request.user.is_superuser

    def handle_no_permission(self):
        if self.request.user.is_authenticated:
            messages.error(self.request, "Access denied. Admin privileges required.")
            return redirect('login')
        return super().handle_no_permission()


class TeacherRequiredMixin(LoginRequiredMixin, UserPassesTestMixin):
    def test_func(self):
        return hasattr(self.request.user, 'teacher_profile')

    def handle_no_permission(self):
        if self.request.user.is_authenticated:
            messages.error(self.request, "Access denied. Teacher privileges required.")
            return redirect('login')
        return super().handle_no_permission()


class StudentRequiredMixin(LoginRequiredMixin, UserPassesTestMixin):
    def test_func(self):
        return hasattr(self.request.user, 'student_profile')

    def handle_no_permission(self):
        if self.request.user.is_authenticated:
            messages.error(self.request, "Access denied. Student privileges required.")
            return redirect('login')
        return super().handle_no_permission()

# ==========================================
# PUBLIC & AUTHENTICATION VIEWS
# ==========================================

class PublicPortalView(TemplateView):
    template_name = 'core/public_portal.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['notices'] = Notice.objects.all().order_by('-created_at')[:8]
        context['courses'] = Course.objects.all()
        return context


class StudentDashboardView(StudentRequiredMixin, View):
    template_name = 'core/student_dashboard.html'

    def get(self, request):
        student = request.user.student_profile
        results = Result.objects.filter(student=student).select_related('subject')
        timetable = Timetable.objects.filter(course=student.course).select_related('subject')
        notices = Notice.objects.filter(Q(subject__course=student.course) | Q(subject__isnull=True)).order_by('-created_at')[:5]

        # Calculate attendance percentage
        att_summary = Attendance.objects.filter(student=student).aggregate(
            total=Count('id'),
            present=Count('id', filter=Q(status='Present'))
        )
        total_att = att_summary['total'] or 0
        present_att = att_summary['present'] or 0
        absent_att = total_att - present_att
        att_pct = (present_att / total_att * 100) if total_att > 0 else 0
        attendance_logs = Attendance.objects.filter(student=student).select_related('subject').order_by('-date')[:15]

        # Calculate GPA / Percentage
        avg_marks = results.aggregate(avg=Avg('marks_obtained'))['avg'] or 0
        avg_max = results.aggregate(avg=Avg('max_marks'))['avg'] or 100
        overall_pct = (avg_marks / avg_max * 100) if results.exists() else 0

        context = {
            'student': student,
            'results': results,
            'timetable': timetable,
            'notices': notices,
            'attendance_percentage': round(att_pct, 1),
            'attendance_total': total_att,
            'attendance_present': present_att,
            'attendance_absent': absent_att,
            'attendance_logs': attendance_logs,
            'overall_percentage': round(overall_pct, 1)
        }
        return render(request, self.template_name, context)


class CampusLoginView(LoginView):
    template_name = 'core/login.html'
    redirect_authenticated_user = True

    def form_valid(self, form):
        user = form.get_user()
        role = self.request.POST.get('role', 'student')
        
        # Check role compatibility
        if role == 'admin' and not user.is_superuser:
            form.add_error(None, "This account does not have Administrator privileges.")
            return self.form_invalid(form)
        elif role == 'staff' and not hasattr(user, 'teacher_profile'):
            form.add_error(None, "This account does not have Staff/Teacher privileges.")
            return self.form_invalid(form)
        elif role == 'student' and not hasattr(user, 'student_profile'):
            form.add_error(None, "This account does not have Student privileges.")
            return self.form_invalid(form)
            
        return super().form_valid(form)


class CampusLogoutView(View):
    def get(self, request, *args, **kwargs):
        from django.contrib.auth import logout
        logout(request)
        return redirect('login')

    def post(self, request, *args, **kwargs):
        from django.contrib.auth import logout
        logout(request)
        return redirect('login')


class CampusPasswordChangeView(LoginRequiredMixin, SuccessMessageMixin, PasswordChangeView):
    form_class = CampusPasswordChangeForm
    template_name = 'core/password_change.html'
    success_url = reverse_lazy('dashboard_router')
    success_message = "Your password has been changed successfully!"


class DashboardRouterView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        if request.user.is_superuser:
            return redirect('admin_dashboard')
        elif hasattr(request.user, 'teacher_profile'):
            return redirect('teacher_dashboard')
        elif hasattr(request.user, 'student_profile'):
            return redirect('student_dashboard')
        else:
            messages.error(request, "User account is misconfigured. Access denied.")
            return redirect('login')

# ==========================================
# ADMIN DASHBOARD & CRUD
# ==========================================

class AdminDashboardView(AdminRequiredMixin, TemplateView):
    template_name = 'core/dashboard_admin.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['total_students'] = Student.objects.count()
        context['total_teachers'] = Teacher.objects.count()
        context['total_courses'] = Course.objects.count()
        context['total_subjects'] = Subject.objects.count()

        # Fee Aggregations
        fee_summary = Fee.objects.aggregate(due=Sum('amount_due'), paid=Sum('amount_paid'))
        due = fee_summary['due'] or 0
        paid = fee_summary['paid'] or 0
        context['fee_due'] = due
        context['fee_paid'] = paid
        context['fee_pending'] = due - paid
        fee_pct = (paid / due * 100) if due > 0 else 0
        context['fee_pct'] = round(fee_pct, 1)

        # Attendance Aggregation
        att_summary = Attendance.objects.aggregate(total=Count('id'), present=Count('id', filter=Q(status='Present')))
        total = att_summary['total'] or 0
        present = att_summary['present'] or 0
        context['attendance_rate'] = round((present / total * 100), 1) if total > 0 else 0.0

        # Recent activities (new additions)
        recent_students = list(Student.objects.all().order_by('-created_at')[:5])
        recent_notices = list(Notice.objects.all().order_by('-created_at')[:5])

        context['recent_students'] = recent_students
        context['recent_notices'] = recent_notices
        context['recent_activities'] = self.generate_recent_activities(recent_students[:3], recent_notices[:3])
        context['recent_fees'] = Fee.objects.select_related('student').all().order_by('-id')[:5]

        return context

    def generate_recent_activities(self, students, notices):
        # Gather dynamic recent occurrences
        activities = []
        for s in students:
            activities.append({
                'title': f"Student Registered: {s.full_name}",
                'time': s.created_at,
                'icon': 'user-plus',
                'bg': 'bg-teal-50 text-teal-600 dark:bg-teal-900/30 dark:text-teal-400'
            })
        for n in notices:
            activities.append({
                'title': f"Announcement: {n.title}",
                'time': n.created_at,
                'icon': 'bell',
                'bg': 'bg-indigo-50 text-indigo-600 dark:bg-indigo-900/30 dark:text-indigo-400'
            })
        activities.sort(key=lambda x: x['time'], reverse=True)
        return activities[:5]


# Student CRUD Views
class AdminStudentListView(AdminRequiredMixin, ListView):
    model = Student
    template_name = 'core/student_list.html'
    context_object_name = 'students'
    paginate_by = 20

    def get_queryset(self):
        queryset = super().get_queryset()
        search_query = self.request.GET.get('search')
        course_filter = self.request.GET.get('course')
        
        if search_query:
            queryset = queryset.filter(
                Q(first_name__icontains=search_query) | 
                Q(last_name__icontains=search_query) | 
                Q(roll_number__icontains=search_query) |
                Q(email__icontains=search_query)
            )
        if course_filter:
            queryset = queryset.filter(course_id=course_filter)
        return queryset.order_by('roll_number')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['courses'] = Course.objects.all()
        return context


class AdminStudentCreateView(AdminRequiredMixin, SuccessMessageMixin, CreateView):
    model = Student
    form_class = StudentForm
    template_name = 'core/student_form.html'
    success_url = reverse_lazy('admin_student_list')
    success_message = "Student profile created successfully."

    def form_valid(self, form):
        student = form.save(commit=False)
        if not student.roll_number:
            student.roll_number = Student.generate_next_roll_number(student.course, student.admission_year)
        username = student.roll_number
        password = form.cleaned_data.get('password') or "Student@123"
        django_user = User.objects.create_user(
            username=username,
            email=student.email,
            first_name=student.first_name,
            last_name=student.last_name,
            password=password
        )
        student.user = django_user
        student.save()
        return super().form_valid(form)


class AdminStudentUpdateView(AdminRequiredMixin, SuccessMessageMixin, UpdateView):
    model = Student
    form_class = StudentForm
    template_name = 'core/student_form.html'
    success_url = reverse_lazy('admin_student_list')
    success_message = "Student profile updated successfully."

    def form_valid(self, form):
        student = form.save(commit=False)
        password = form.cleaned_data.get('password')
        if student.user:
            django_user = student.user
            django_user.username = student.roll_number
            django_user.email = student.email
            django_user.first_name = student.first_name
            django_user.last_name = student.last_name
            if password:
                django_user.set_password(password)
            django_user.save()
        else:
            if not student.roll_number:
                student.roll_number = Student.generate_next_roll_number(student.course, student.admission_year)
            username = student.roll_number
            default_pwd = password or "Student@123"
            django_user = User.objects.create_user(
                username=username,
                email=student.email,
                first_name=student.first_name,
                last_name=student.last_name,
                password=default_pwd
            )
            student.user = django_user
        student.save()
        return super().form_valid(form)


class AdminStudentDeleteView(AdminRequiredMixin, DeleteView):
    model = Student
    template_name = 'core/confirm_delete.html'
    success_url = reverse_lazy('admin_student_list')

    def post(self, request, *args, **kwargs):
        messages.success(request, "Student profile deleted successfully.")
        return super().post(request, *args, **kwargs)


# Teacher CRUD Views
class AdminTeacherListView(AdminRequiredMixin, ListView):
    model = Teacher
    template_name = 'core/teacher_list.html'
    context_object_name = 'teachers'
    paginate_by = 20

    def get_queryset(self):
        queryset = super().get_queryset()
        search_query = self.request.GET.get('search')
        if search_query:
            queryset = queryset.filter(
                Q(user__first_name__icontains=search_query) | 
                Q(user__last_name__icontains=search_query) | 
                Q(qualification__icontains=search_query) | 
                Q(phone__icontains=search_query)
            )
        return queryset.order_by('user__first_name')


class AdminTeacherCreateView(AdminRequiredMixin, View):
    template_name = 'core/teacher_form.html'

    def get(self, request):
        user_form = TeacherUserForm()
        profile_form = TeacherProfileForm()
        return render(request, self.template_name, {
            'user_form': user_form,
            'profile_form': profile_form,
            'title': 'Add Teacher'
        })

    def post(self, request):
        user_form = TeacherUserForm(request.POST)
        profile_form = TeacherProfileForm(request.POST, request.FILES)
        if user_form.is_valid() and profile_form.is_valid():
            user = user_form.save(commit=False)
            password = user_form.cleaned_data.get('password')
            if not password:
                password = 'Teacher@123'
            user.set_password(password)
            user.save()
            
            profile = profile_form.save(commit=False)
            profile.user = user
            profile.save()
            messages.success(request, "Teacher registered successfully!")
            return redirect('admin_teacher_list')
        
        return render(request, self.template_name, {
            'user_form': user_form,
            'profile_form': profile_form,
            'title': 'Add Teacher'
        })


class AdminTeacherUpdateView(AdminRequiredMixin, View):
    template_name = 'core/teacher_form.html'

    def get(self, request, pk):
        teacher = get_object_or_404(Teacher, pk=pk)
        user_form = TeacherUserForm(instance=teacher.user)
        profile_form = TeacherProfileForm(instance=teacher)
        return render(request, self.template_name, {
            'user_form': user_form,
            'profile_form': profile_form,
            'title': 'Edit Teacher'
        })

    def post(self, request, pk):
        teacher = get_object_or_404(Teacher, pk=pk)
        user_form = TeacherUserForm(request.POST, instance=teacher.user)
        profile_form = TeacherProfileForm(request.POST, request.FILES, instance=teacher)
        if user_form.is_valid() and profile_form.is_valid():
            user = user_form.save(commit=False)
            password = user_form.cleaned_data.get('password')
            if password:
                user.set_password(password)
            user.save()
            profile_form.save()
            messages.success(request, "Teacher details updated successfully.")
            return redirect('admin_teacher_list')
        
        return render(request, self.template_name, {
            'user_form': user_form,
            'profile_form': profile_form,
            'title': 'Edit Teacher'
        })


class AdminTeacherDeleteView(AdminRequiredMixin, View):
    def post(self, request, pk):
        teacher = get_object_or_404(Teacher, pk=pk)
        user = teacher.user
        teacher.delete()
        user.delete()
        messages.success(request, "Teacher deleted successfully.")
        return redirect('admin_teacher_list')


# Course CRUD
class AdminCourseListView(AdminRequiredMixin, View):
    template_name = 'core/course_list.html'

    def get(self, request):
        courses = Course.objects.all().order_by('name')
        form = CourseForm()
        return render(request, self.template_name, {'courses': courses, 'form': form})

    def post(self, request):
        courses = Course.objects.all().order_by('name')
        form = CourseForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Course created successfully.")
            return redirect('admin_course_list')
        return render(request, self.template_name, {'courses': courses, 'form': form})


class AdminCourseUpdateView(AdminRequiredMixin, SuccessMessageMixin, UpdateView):
    model = Course
    form_class = CourseForm
    template_name = 'core/course_form.html'
    success_url = reverse_lazy('admin_course_list')
    success_message = "Course updated successfully."


class AdminCourseDeleteView(AdminRequiredMixin, DeleteView):
    model = Course
    template_name = 'core/confirm_delete.html'
    success_url = reverse_lazy('admin_course_list')

    def post(self, request, *args, **kwargs):
        messages.success(request, "Course deleted successfully.")
        return super().post(request, *args, **kwargs)


# Subject CRUD
class AdminSubjectListView(AdminRequiredMixin, View):
    template_name = 'core/subject_list.html'

    def get(self, request):
        subjects = Subject.objects.all().select_related('course', 'teacher').order_by('name')
        form = SubjectForm()
        return render(request, self.template_name, {'subjects': subjects, 'form': form})

    def post(self, request):
        subjects = Subject.objects.all().select_related('course', 'teacher').order_by('name')
        form = SubjectForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Subject created successfully.")
            return redirect('admin_subject_list')
        return render(request, self.template_name, {'subjects': subjects, 'form': form})


class AdminSubjectUpdateView(AdminRequiredMixin, SuccessMessageMixin, UpdateView):
    model = Subject
    form_class = SubjectForm
    template_name = 'core/subject_form.html'
    success_url = reverse_lazy('admin_subject_list')
    success_message = "Subject updated successfully."


class AdminSubjectDeleteView(AdminRequiredMixin, DeleteView):
    model = Subject
    template_name = 'core/confirm_delete.html'
    success_url = reverse_lazy('admin_subject_list')

    def post(self, request, *args, **kwargs):
        messages.success(request, "Subject deleted successfully.")
        return super().post(request, *args, **kwargs)


# Attendance Management
class AdminAttendanceListView(AdminRequiredMixin, ListView):
    model = Attendance
    template_name = 'core/admin_attendance_list.html'
    context_object_name = 'attendance_records'
    paginate_by = 30
    ordering = ['-date']

    def get_queryset(self):
        qs = super().get_queryset()
        student_search = self.request.GET.get('search')
        date_filter = self.request.GET.get('date')
        if student_search:
            qs = qs.filter(student__first_name__icontains=student_search) | qs.filter(student__last_name__icontains=student_search)
        if date_filter:
            qs = qs.filter(date=date_filter)
        return qs.select_related('student', 'subject')


class AdminAttendanceMarkView(AdminRequiredMixin, View):
    template_name = 'core/admin_attendance_mark.html'

    def get(self, request):
        subjects = Subject.objects.all()
        students = []
        subject_id = request.GET.get('subject')
        date_val = request.GET.get('date', datetime.date.today().strftime('%Y-%m-%d'))
        
        selected_subject = None
        if subject_id:
            selected_subject = get_object_or_404(Subject, pk=subject_id)
            raw_students = Student.objects.filter(course=selected_subject.course).order_by('first_name', 'last_name')
            existing = Attendance.objects.filter(subject=selected_subject, date=date_val)
            existing_map = {att.student_id: att.status for att in existing}
            for s in raw_students:
                s.existing_status = existing_map.get(s.id, 'Present')
                students.append(s)

        return render(request, self.template_name, {
            'subjects': subjects,
            'students': students,
            'date': date_val,
            'selected_subject': selected_subject
        })

    def post(self, request):
        subject_id = request.POST.get('subject_id')
        date_val = request.POST.get('date')
        if not subject_id or not date_val:
            messages.error(request, "Invalid details provided.")
            return redirect('admin_attendance_mark')
            
        subject = get_object_or_404(Subject, pk=subject_id)
        students = Student.objects.filter(course=subject.course)
        
        for student in students:
            status = request.POST.get(f'status_{student.id}')
            if status in ['Present', 'Absent']:
                Attendance.objects.update_or_create(
                    student=student,
                    subject=subject,
                    date=date_val,
                    defaults={'status': status, 'marked_by': request.user}
                )
        messages.success(request, f"Attendance updated successfully for {subject.name} on {date_val}.")
        return redirect('admin_attendance_list')


# Result Management
class AdminResultListView(AdminRequiredMixin, ListView):
    model = Result
    template_name = 'core/admin_result_list.html'
    context_object_name = 'results'
    paginate_by = 30
    ordering = ['-exam_date']

    def get_queryset(self):
        qs = super().get_queryset()
        search_query = self.request.GET.get('search')
        if search_query:
            qs = qs.filter(student__first_name__icontains=search_query) | qs.filter(student__last_name__icontains=search_query)
        return qs.select_related('student', 'subject')


class AdminResultCreateView(AdminRequiredMixin, SuccessMessageMixin, CreateView):
    model = Result
    form_class = ResultForm
    template_name = 'core/admin_result_form.html'
    success_url = reverse_lazy('admin_result_list')
    success_message = "Result created successfully."

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        # Auto compute grade based on marks
        marks = form.cleaned_data.get('marks_obtained')
        max_marks = form.cleaned_data.get('max_marks', 100.0)
        percentage = (marks / max_marks) * 100
        if percentage >= 90: form.instance.grade = 'A+'
        elif percentage >= 80: form.instance.grade = 'A'
        elif percentage >= 70: form.instance.grade = 'B'
        elif percentage >= 60: form.instance.grade = 'C'
        elif percentage >= 50: form.instance.grade = 'D'
        elif percentage >= 40: form.instance.grade = 'E'
        else: form.instance.grade = 'F'
        return super().form_valid(form)


class AdminResultUpdateView(AdminRequiredMixin, SuccessMessageMixin, UpdateView):
    model = Result
    form_class = ResultForm
    template_name = 'core/admin_result_form.html'
    success_url = reverse_lazy('admin_result_list')
    success_message = "Result updated successfully."

    def form_valid(self, form):
        marks = form.cleaned_data.get('marks_obtained')
        max_marks = form.cleaned_data.get('max_marks', 100.0)
        percentage = (marks / max_marks) * 100
        if percentage >= 90: form.instance.grade = 'A+'
        elif percentage >= 80: form.instance.grade = 'A'
        elif percentage >= 70: form.instance.grade = 'B'
        elif percentage >= 60: form.instance.grade = 'C'
        elif percentage >= 50: form.instance.grade = 'D'
        elif percentage >= 40: form.instance.grade = 'E'
        else: form.instance.grade = 'F'
        return super().form_valid(form)


class AdminResultDeleteView(AdminRequiredMixin, DeleteView):
    model = Result
    template_name = 'core/confirm_delete.html'
    success_url = reverse_lazy('admin_result_list')

    def post(self, request, *args, **kwargs):
        messages.success(request, "Result deleted successfully.")
        return super().post(request, *args, **kwargs)


# Fee Management
class AdminFeeListView(AdminRequiredMixin, ListView):
    model = Fee
    template_name = 'core/admin_fee_list.html'
    context_object_name = 'fees'
    paginate_by = 30

    def get_queryset(self):
        qs = super().get_queryset()
        search_query = self.request.GET.get('search')
        status_filter = self.request.GET.get('status')
        if search_query:
            qs = qs.filter(student__first_name__icontains=search_query) | qs.filter(student__last_name__icontains=search_query)
        if status_filter:
            qs = qs.filter(status=status_filter)
        return qs.select_related('student')


class AdminFeeCreateView(AdminRequiredMixin, SuccessMessageMixin, CreateView):
    model = Fee
    form_class = FeeForm
    template_name = 'core/admin_fee_form.html'
    success_url = reverse_lazy('admin_fee_list')
    success_message = "Fee record created successfully."


class AdminFeeUpdateView(AdminRequiredMixin, SuccessMessageMixin, UpdateView):
    model = Fee
    form_class = FeeForm
    template_name = 'core/admin_fee_form.html'
    success_url = reverse_lazy('admin_fee_list')
    success_message = "Fee record updated successfully."


class AdminFeeDeleteView(AdminRequiredMixin, DeleteView):
    model = Fee
    template_name = 'core/confirm_delete.html'
    success_url = reverse_lazy('admin_fee_list')

    def post(self, request, *args, **kwargs):
        messages.success(request, "Fee record deleted successfully.")
        return super().post(request, *args, **kwargs)


# Timetable Management
class AdminTimetableListView(AdminRequiredMixin, ListView):
    model = Timetable
    template_name = 'core/admin_timetable_list.html'
    context_object_name = 'slots'
    ordering = ['course', 'day', 'start_time']

    def get_queryset(self):
        qs = super().get_queryset()
        course_filter = self.request.GET.get('course')
        if course_filter:
            qs = qs.filter(course_id=course_filter)
        return qs.select_related('course', 'subject')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['courses'] = Course.objects.all()
        return context


class AdminTimetableCreateView(AdminRequiredMixin, SuccessMessageMixin, CreateView):
    model = Timetable
    form_class = TimetableForm
    template_name = 'core/admin_timetable_form.html'
    success_url = reverse_lazy('admin_timetable_list')
    success_message = "Timetable slot added successfully."


class AdminTimetableUpdateView(AdminRequiredMixin, SuccessMessageMixin, UpdateView):
    model = Timetable
    form_class = TimetableForm
    template_name = 'core/admin_timetable_form.html'
    success_url = reverse_lazy('admin_timetable_list')
    success_message = "Timetable slot updated successfully."


class AdminTimetableDeleteView(AdminRequiredMixin, DeleteView):
    model = Timetable
    template_name = 'core/confirm_delete.html'
    success_url = reverse_lazy('admin_timetable_list')

    def post(self, request, *args, **kwargs):
        messages.success(request, "Timetable slot deleted successfully.")
        return super().post(request, *args, **kwargs)


# Notice Management
class AdminNoticeListView(AdminRequiredMixin, ListView):
    model = Notice
    template_name = 'core/admin_notice_list.html'
    context_object_name = 'notices'
    ordering = ['-created_at']


class AdminNoticeCreateView(AdminRequiredMixin, SuccessMessageMixin, CreateView):
    model = Notice
    form_class = NoticeForm
    template_name = 'core/admin_notice_form.html'
    success_url = reverse_lazy('admin_notice_list')
    success_message = "Announcement posted successfully."

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        return super().form_valid(form)


class AdminNoticeUpdateView(AdminRequiredMixin, SuccessMessageMixin, UpdateView):
    model = Notice
    form_class = NoticeForm
    template_name = 'core/admin_notice_form.html'
    success_url = reverse_lazy('admin_notice_list')
    success_message = "Announcement updated successfully."


class AdminNoticeDeleteView(AdminRequiredMixin, DeleteView):
    model = Notice
    template_name = 'core/confirm_delete.html'
    success_url = reverse_lazy('admin_notice_list')

    def post(self, request, *args, **kwargs):
        messages.success(request, "Announcement deleted successfully.")
        return super().post(request, *args, **kwargs)

# ==========================================
# TEACHER DASHBOARD & FEATURES
# ==========================================

class TeacherDashboardView(TeacherRequiredMixin, TemplateView):
    template_name = 'core/dashboard_teacher.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        teacher = self.request.user.teacher_profile
        subjects = Subject.objects.filter(teacher=teacher)
        
        # Aggregate statistics
        context['teacher'] = teacher
        context['subjects'] = subjects
        context['assigned_subjects_count'] = subjects.count()
        
        courses_taught = Course.objects.filter(subjects__teacher=teacher).distinct()
        students_count = Student.objects.filter(course__in=courses_taught).count()
        context['students_count'] = students_count

        # Notices posted by this teacher
        context['notices'] = Notice.objects.filter(created_by=self.request.user).order_by('-created_at')[:5]
        
        return context


class TeacherSubjectListView(TeacherRequiredMixin, ListView):
    model = Subject
    template_name = 'core/teacher_subject_list.html'
    context_object_name = 'subjects'

    def get_queryset(self):
        return Subject.objects.filter(teacher=self.request.user.teacher_profile).select_related('course')


class TeacherAttendanceMarkView(TeacherRequiredMixin, View):
    template_name = 'core/teacher_attendance_mark.html'

    def get(self, request, subject_id):
        subject = get_object_or_404(Subject, pk=subject_id, teacher=request.user.teacher_profile)
        raw_students = Student.objects.filter(course=subject.course).order_by('first_name', 'last_name')
        date_str = request.GET.get('date', datetime.date.today().strftime('%Y-%m-%d'))
        
        # Fetch existing markings
        existing = Attendance.objects.filter(subject=subject, date=date_str)
        existing_map = {att.student_id: att.status for att in existing}
        
        students = []
        for s in raw_students:
            s.existing_status = existing_map.get(s.id, 'Present')
            students.append(s)
        
        return render(request, self.template_name, {
            'subject': subject,
            'students': students,
            'date': date_str
        })

    def post(self, request, subject_id):
        subject = get_object_or_404(Subject, pk=subject_id, teacher=request.user.teacher_profile)
        students = Student.objects.filter(course=subject.course)
        date_val = request.POST.get('date')
        if not date_val:
            date_val = datetime.date.today().strftime('%Y-%m-%d')
            
        for student in students:
            status = request.POST.get(f'status_{student.id}')
            if status in ['Present', 'Absent']:
                Attendance.objects.update_or_create(
                    student=student,
                    subject=subject,
                    date=date_val,
                    defaults={'status': status, 'marked_by': request.user}
                )
        messages.success(request, f"Attendance for {subject.name} on {date_val} updated successfully.")
        return redirect('teacher_dashboard')


class TeacherResultManageView(TeacherRequiredMixin, View):
    template_name = 'core/teacher_result_manage.html'

    def get(self, request, subject_id):
        subject = get_object_or_404(Subject, pk=subject_id, teacher=request.user.teacher_profile)
        raw_students = Student.objects.filter(course=subject.course).order_by('first_name', 'last_name')
        results = Result.objects.filter(subject=subject)
        results_map = {res.student_id: res for res in results}
        
        students = []
        for s in raw_students:
            res = results_map.get(s.id)
            s.existing_marks = res.marks_obtained if res else ''
            s.existing_max_marks = res.max_marks if res else 100.00
            students.append(s)
        
        return render(request, self.template_name, {
            'subject': subject,
            'students': students,
            'default_date': datetime.date.today().strftime('%Y-%m-%d')
        })

    def post(self, request, subject_id):
        subject = get_object_or_404(Subject, pk=subject_id, teacher=request.user.teacher_profile)
        students = Student.objects.filter(course=subject.course)
        exam_date_val = request.POST.get('exam_date')
        if not exam_date_val:
            exam_date_val = datetime.date.today().strftime('%Y-%m-%d')
            
        for student in students:
            marks_str = request.POST.get(f'marks_{student.id}')
            max_marks_str = request.POST.get(f'max_marks_{student.id}', '100')
            if marks_str:
                try:
                    marks = float(marks_str)
                    max_marks = float(max_marks_str)
                    if marks > max_marks:
                        messages.error(request, f"Marks for {student.full_name} cannot exceed max marks.")
                        continue
                    
                    # Compute Grade
                    pct = (marks / max_marks) * 100
                    if pct >= 90: grade = 'A+'
                    elif pct >= 80: grade = 'A'
                    elif pct >= 70: grade = 'B'
                    elif pct >= 60: grade = 'C'
                    elif pct >= 50: grade = 'D'
                    elif pct >= 40: grade = 'E'
                    else: grade = 'F'
                    
                    Result.objects.update_or_create(
                        student=student,
                        subject=subject,
                        exam_date=exam_date_val,
                        defaults={
                            'marks_obtained': marks,
                            'max_marks': max_marks,
                            'grade': grade,
                            'created_by': request.user
                        }
                    )
                except ValueError:
                    pass
        messages.success(request, f"Exam results updated successfully for {subject.name}.")
        return redirect('teacher_dashboard')


class TeacherStudentListView(TeacherRequiredMixin, ListView):
    model = Student
    template_name = 'core/teacher_student_list.html'
    context_object_name = 'students'

    def get_queryset(self):
        teacher = self.request.user.teacher_profile
        courses_taught = Course.objects.filter(subjects__teacher=teacher).distinct()
        
        qs = Student.objects.filter(course__in=courses_taught).select_related('course')
        search_query = self.request.GET.get('search')
        if search_query:
            qs = qs.filter(
                Q(first_name__icontains=search_query) | 
                Q(last_name__icontains=search_query) | 
                Q(roll_number__icontains=search_query)
            )
        return qs.order_by('roll_number')


class TeacherNoticeListView(TeacherRequiredMixin, ListView):
    model = Notice
    template_name = 'core/teacher_notice_list.html'
    context_object_name = 'notices'

    def get_queryset(self):
        return Notice.objects.filter(created_by=self.request.user).order_by('-created_at')


class TeacherNoticeCreateView(TeacherRequiredMixin, SuccessMessageMixin, CreateView):
    model = Notice
    form_class = NoticeForm
    template_name = 'core/teacher_notice_form.html'
    success_url = reverse_lazy('teacher_notice_list')
    success_message = "Subject announcement published."

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        return super().form_valid(form)

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        # Limit options in subject selection to only subjects taught by the teacher
        form.fields['subject'].queryset = Subject.objects.filter(teacher=self.request.user.teacher_profile)
        form.fields['subject'].required = True
        return form

# ==========================================
# REPORTS & EXPORTS
# ==========================================

class AdminReportsView(AdminRequiredMixin, TemplateView):
    template_name = 'core/admin_reports.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['students'] = Student.objects.all().order_by('roll_number')
        return context


@admin_required
def export_students_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Students Report"

    # Define layout designs
    title_font = Font(name="Segoe UI", size=16, bold=True, color="FFFFFF")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Segoe UI", size=10)
    
    title_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid") # Dark Blue
    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid") # Light Blue
    alt_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid") # Zebra gray

    border_thin = Border(
        left=Side(style='thin', color="E5E7EB"),
        right=Side(style='thin', color="E5E7EB"),
        top=Side(style='thin', color="E5E7EB"),
        bottom=Side(style='thin', color="E5E7EB")
    )

    # Title Banner
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = "CampusCore - Student Enrollment Directory"
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    # Table Header
    headers = ["Roll Number", "First Name", "Last Name", "Email Address", "Phone Number", "Gender", "Date of Birth", "Enrolled Course"]
    ws.append([]) # Empty spacing row
    ws.append(headers)
    
    ws.row_dimensions[3].height = 25
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=3, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_thin

    # Append Data
    students = Student.objects.select_related('course').all().order_by('roll_number')
    row_num = 4
    for idx, student in enumerate(students):
        row_data = [
            student.roll_number,
            student.first_name,
            student.last_name,
            student.email,
            student.phone,
            student.gender,
            student.dob.strftime('%Y-%m-%d') if student.dob else '',
            student.course.name
        ]
        ws.append(row_data)
        ws.row_dimensions[row_num].height = 20
        
        # Styling cells
        for col_idx in range(1, len(row_data) + 1):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.font = data_font
            cell.border = border_thin
            if idx % 2 == 1:
                cell.fill = alt_fill
            # Alignments
            if col_idx in [1, 5, 6, 7]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
        row_num += 1

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.row == 1: continue # Skip title row for calculation
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="CampusCore_Students_List.xlsx"'
    wb.save(response)
    return response


@admin_required
def export_teachers_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Teachers Directory"

    title_font = Font(name="Segoe UI", size=16, bold=True, color="FFFFFF")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Segoe UI", size=10)
    
    title_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid") # Dark Slate
    header_fill = PatternFill(start_color="475569", end_color="475569", fill_type="solid") # Slate
    alt_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

    border_thin = Border(
        left=Side(style='thin', color="E2E8F0"),
        right=Side(style='thin', color="E2E8F0"),
        top=Side(style='thin', color="E2E8F0"),
        bottom=Side(style='thin', color="E2E8F0")
    )

    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = "CampusCore - Teacher Faculty Directory"
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    headers = ["Username", "Full Name", "Email Address", "Phone Number", "Qualification", "Office Address"]
    ws.append([]) # Space
    ws.append(headers)
    
    ws.row_dimensions[3].height = 25
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=3, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_thin

    teachers = Teacher.objects.select_related('user').all().order_by('user__first_name')
    row_num = 4
    for idx, teacher in enumerate(teachers):
        row_data = [
            teacher.user.username,
            teacher.full_name,
            teacher.user.email,
            teacher.phone,
            teacher.qualification,
            teacher.address
        ]
        ws.append(row_data)
        ws.row_dimensions[row_num].height = 20
        
        for col_idx in range(1, len(row_data) + 1):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.font = data_font
            cell.border = border_thin
            if idx % 2 == 1:
                cell.fill = alt_fill
            if col_idx in [1, 4]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
        row_num += 1

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.row == 1: continue
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="CampusCore_Teachers_List.xlsx"'
    wb.save(response)
    return response


@admin_required
def export_student_pdf(request, student_id):
    student = get_object_or_404(Student, pk=student_id)
    results = Result.objects.filter(student=student).select_related('subject')
    
    # Calculate Attendance stats
    att_records = Attendance.objects.filter(student=student)
    total_att = att_records.count()
    present_att = att_records.filter(status='Present').count()
    absent_att = total_att - present_att
    attendance_pct = (present_att / total_att * 100) if total_att > 0 else 0.0

    # Document Template Setup
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="CampusCore_ReportCard_{student.roll_number}.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=letter, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
    story = []
    
    styles = getSampleStyleSheet()
    
    # Custom Styles
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=24,
        textColor=colors.HexColor('#1E3A8A'),
        alignment=1, # Center
        spaceAfter=5
    )
    slogan_style = ParagraphStyle(
        'DocSlogan',
        parent=styles['Normal'],
        fontName='Helvetica-Oblique',
        fontSize=10,
        textColor=colors.HexColor('#6B7280'),
        alignment=1,
        spaceAfter=15
    )
    section_heading = ParagraphStyle(
        'SectionHeading',
        parent=styles['Heading2'],
        fontName='Helvetica-Bold',
        fontSize=14,
        textColor=colors.HexColor('#1E3A8A'),
        spaceBefore=15,
        spaceAfter=8,
        borderPadding=(0,0,2,0)
    )
    body_style = ParagraphStyle(
        'CardBody',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=10,
        textColor=colors.HexColor('#111827')
    )
    body_bold = ParagraphStyle(
        'CardBodyBold',
        parent=body_style,
        fontName='Helvetica-Bold'
    )
    
    # Title Block
    story.append(Paragraph("CampusCore", title_style))
    story.append(Paragraph("The Core of Campus Management &mdash; Performance & Attendance Summary", slogan_style))
    story.append(Spacer(1, 10))
    
    # Info Grid (Student Metadata)
    metadata = [
        [Paragraph("<b>Student Name:</b>", body_style), Paragraph(student.full_name, body_style),
         Paragraph("<b>Roll Number:</b>", body_style), Paragraph(student.roll_number, body_style)],
        [Paragraph("<b>Enrolled Course:</b>", body_style), Paragraph(student.course.name, body_style),
         Paragraph("<b>Date of Birth:</b>", body_style), Paragraph(student.dob.strftime('%B %d, %Y') if student.dob else 'N/A', body_style)],
        [Paragraph("<b>Email:</b>", body_style), Paragraph(student.email, body_style),
         Paragraph("<b>Contact:</b>", body_style), Paragraph(student.phone, body_style)]
    ]
    meta_table = Table(metadata, colWidths=[110, 160, 100, 160])
    meta_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('LINEBELOW', (0,-1), (-1,-1), 1, colors.HexColor('#E5E7EB')),
    ]))
    story.append(meta_table)
    story.append(Spacer(1, 15))
    
    # Section: Results
    story.append(Paragraph("Academic Performance Card", section_heading))
    
    res_data = [[
        Paragraph("<b>Subject Code</b>", body_bold),
        Paragraph("<b>Subject Name</b>", body_bold),
        Paragraph("<b>Marks Obtained</b>", body_bold),
        Paragraph("<b>Max Marks</b>", body_bold),
        Paragraph("<b>Grade</b>", body_bold)
    ]]
    
    total_marks = 0
    total_max = 0
    for res in results:
        res_data.append([
            Paragraph(res.subject.code, body_style),
            Paragraph(res.subject.name, body_style),
            Paragraph(str(res.marks_obtained), body_style),
            Paragraph(str(res.max_marks), body_style),
            Paragraph(res.grade, body_bold)
        ])
        total_marks += res.marks_obtained
        total_max += res.max_marks

    # Compute Summary
    avg_percentage = (total_marks / total_max * 100) if total_max > 0 else 0.0
    
    res_table = Table(res_data, colWidths=[90, 190, 90, 90, 70])
    res_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#F3F4F6')),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#D1D5DB')),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
    ]))
    story.append(res_table)
    
    # Academic Summary Stats
    story.append(Spacer(1, 10))
    summary_txt = f"<b>Total Aggregated Marks:</b> {total_marks} / {total_max} &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp; <b>Cumulative Percentage:</b> {round(avg_percentage, 2)}%"
    story.append(Paragraph(summary_txt, body_style))
    story.append(Spacer(1, 15))
    
    # Section: Attendance
    story.append(Paragraph("Attendance Log Summary", section_heading))
    att_data = [
        [Paragraph("<b>Status</b>", body_bold), Paragraph("<b>Count</b>", body_bold), Paragraph("<b>Percentage Representation</b>", body_bold)],
        [Paragraph("Days Present", body_style), Paragraph(str(present_att), body_style), Paragraph(f"{round(attendance_pct, 1)}%", body_bold)],
        [Paragraph("Days Absent", body_style), Paragraph(str(absent_att), body_style), Paragraph(f"{round(100 - attendance_pct, 1) if total_att > 0 else 0.0}%", body_style)],
        [Paragraph("Total Academic Days", body_bold), Paragraph(str(total_att), body_bold), Paragraph("100.0%", body_bold)]
    ]
    att_table = Table(att_data, colWidths=[180, 150, 200])
    att_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#F3F4F6')),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#D1D5DB')),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor('#F9FAFB')),
    ]))
    story.append(att_table)
    story.append(Spacer(1, 40))
    
    # Signatures
    sig_data = [
        [Paragraph("_____________________________<br/><b>Faculty Coordinator</b>", body_style),
         Paragraph("_____________________________<br/><b>Principal / Registrar Office</b>", body_style)]
    ]
    sig_table = Table(sig_data, colWidths=[270, 260])
    sig_table.setStyle(TableStyle([
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
    ]))
    story.append(sig_table)

    doc.build(story)
    return response
